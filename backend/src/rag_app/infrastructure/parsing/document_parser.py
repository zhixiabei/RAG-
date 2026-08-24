from __future__ import annotations

import csv
from html.parser import HTMLParser
from io import BytesIO, StringIO
import json
import os
from pathlib import Path
import posixpath
import re
import struct
import sys
import tempfile
from typing import Any, BinaryIO, Iterable
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile

from ...domain.models import ParsedChunk


SUPPORTED_SUFFIXES = frozenset(
    {
        ".jsonl",
        ".json",
        ".pdf",
        ".doc",
        ".docx",
        ".pptx",
        ".xlsx",
        ".md",
        ".markdown",
        ".txt",
    }
)

PARSABLE_SUFFIXES = SUPPORTED_SUFFIXES | frozenset(
    {
        ".doc",
        ".xlsm",
        ".xls",
        ".csv",
        ".html",
        ".htm",
        ".xml",
        ".dll",
        ".gdb",
        ".att",
        ".ptpt",
        ".jcpt",
        ".stpt",
        ".ppt",
        ".lst",
        "",
    }
)

CHUNK_SIZE = 400
CHUNK_OVERLAP = 50
MIN_NATURAL_CHUNK_SIZE = CHUNK_SIZE // 2

MARKDOWN_HEADING_PATTERN = re.compile(r"^(#{1,6})\s+(.+?)(?:\s+#+)?$")
NUMBERED_HEADING_PATTERNS = (
    (re.compile(r"^第[一二三四五六七八九十百千万零〇两\d]+[篇章部]\s*[:：]?\s*(.*)$"), 1),
    (re.compile(r"^第[一二三四五六七八九十百千万零〇两\d]+节\s*[:：]?\s*(.*)$"), 2),
    (re.compile(r"^第[一二三四五六七八九十百千万零〇两\d]+条\s*[:：]?\s*(.*)$"), 3),
    (re.compile(r"^[一二三四五六七八九十百千万零〇两]+、\s*(.+)$"), 1),
    (re.compile(r"^[（(][一二三四五六七八九十百千万零〇两]+[）)]\s*(.+)$"), 2),
    (re.compile(r"^(\d+(?:\.\d+)+)[.、]?\s+(.+)$"), None),
    (re.compile(r"^\d+[.、]\s*(.+)$"), 1),
)

GEOMAP_MAP_MAGIC = b"Geomap v3.60 Map\x00"
GEOMAP_LAYER_MAGIC = b"Geomap v3.60 Layer\x00\x00"
GEOMAP_LAYER_HEADER_SIZE = 2048
GEOMAP_LAYER_NAME_LENGTH_OFFSET = 0x362
GEOMAP_LAYER_NAME_OFFSET = 0x366
GEOMAP_OBJECT_TEXT_MARKER = b"\x2f\x3d\x50\x75"
GEOMAP_MAX_TEXT_FIELD_BYTES = 1024
GEOMAP_MAX_ATTRIBUTE_ROWS = 100_000
GEOMAP_LAYER_STYLE_MAGIC = b"Geomap v3.60 LayerStyle\x00"
GEOMAP_LAYER_STYLE_HEADER_SIZE = 516
GEOMAP_LAYER_STYLE_RECORD_SIZE = 524
GEOMAP_LAYER_STYLE_TEXT_SLOTS = (
    (0, 64),
    (80, 64),
    (160, 64),
    (240, 64),
    (320, 64),
    (400, 64),
    (480, 32),
)
GEOMAP_ALBUM_MAGIC = b"GeoMap Album Information\n"
GEOMAP_ALBUM_TREE_OFFSET = 2048
PACKAGE_RELATIONSHIPS_NAMESPACE = "http://schemas.openxmlformats.org/package/2006/relationships"
OFFICE_RELATIONSHIPS_NAMESPACE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PRESENTATIONML_NAMESPACE = "http://schemas.openxmlformats.org/presentationml/2006/main"
DRAWINGML_NAMESPACE = "http://schemas.openxmlformats.org/drawingml/2006/main"
PPTX_MAX_SLIDE_XML_BYTES = 32 * 1024 * 1024
OPTIONAL_OFFICE_UI_PART_PREFIXES = ("customUI/", "userCustomization/")


class UnsupportedDocumentTypeError(ValueError):
    pass


class _TextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.parts: list[str] = []

    def handle_data(self, data: str) -> None:
        if data.strip():
            self.parts.append(data.strip())


def _decode_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "utf-16"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _relationship_source_directory(relationship_file: str) -> str:
    directory, file_name = posixpath.split(relationship_file)
    if posixpath.basename(directory) != "_rels" or not file_name.endswith(".rels"):
        return ""
    source_directory = posixpath.dirname(directory)
    return source_directory


def _repair_missing_optional_docx_parts(content: bytes) -> bytes:
    """Remove broken relationships to optional Office UI customization parts."""
    with ZipFile(BytesIO(content)) as source:
        archive_names = set(source.namelist())
        rewritten_relationships: dict[str, bytes] = {}
        relationship_tag = f"{{{PACKAGE_RELATIONSHIPS_NAMESPACE}}}Relationship"

        for entry in source.infolist():
            if not entry.filename.endswith(".rels"):
                continue
            relationship_xml = source.read(entry)
            try:
                root = ElementTree.fromstring(relationship_xml)
            except ElementTree.ParseError:
                continue

            source_directory = _relationship_source_directory(entry.filename)
            changed = False
            for relationship in list(root):
                if relationship.tag != relationship_tag or relationship.get("TargetMode") == "External":
                    continue
                target = relationship.get("Target", "").replace("\\", "/")
                target_part = posixpath.normpath(
                    target.lstrip("/") if target.startswith("/") else posixpath.join(source_directory, target)
                )
                is_optional_ui_part = target_part.startswith(OPTIONAL_OFFICE_UI_PART_PREFIXES)
                if is_optional_ui_part and target_part not in archive_names:
                    root.remove(relationship)
                    changed = True

            if changed:
                rewritten_relationships[entry.filename] = ElementTree.tostring(
                    root,
                    encoding="utf-8",
                    xml_declaration=True,
                )

        if not rewritten_relationships:
            return content

        repaired = BytesIO()
        with ZipFile(repaired, "w") as destination:
            for entry in source.infolist():
                destination.writestr(
                    entry,
                    rewritten_relationships.get(entry.filename, source.read(entry)),
                )
        return repaired.getvalue()


class DocumentParser:
    def supports(self, file_name: str) -> bool:
        return Path(file_name).suffix.lower() in SUPPORTED_SUFFIXES

    def parse(self, file_name: str, content: bytes) -> list[ParsedChunk]:
        return self.parse_stream(file_name, BytesIO(content))

    def parse_stream(self, file_name: str, stream: BinaryIO) -> list[ParsedChunk]:
        suffix = Path(file_name).suffix.lower()
        if suffix not in PARSABLE_SUFFIXES:
            raise UnsupportedDocumentTypeError(f"暂不支持的文件类型: {suffix or '无扩展名'}")

        stream.seek(0)
        if suffix == ".pdf":
            sources = self._parse_pdf_stream(stream)
        elif suffix == ".docx":
            sources = self._parse_docx_stream(stream)
        elif suffix == ".doc":
            sources = self._parse_doc(stream.read())
        elif suffix == ".pptx":
            sources = self._parse_pptx_stream(stream)
        elif suffix in {".xlsx", ".xlsm"}:
            sources = self._parse_xlsx_stream(stream)
        elif suffix == ".xls":
            sources = self._parse_xls(stream.read())
        elif suffix == ".csv":
            sources = self._parse_csv(stream.read())
        elif suffix == ".jsonl":
            sources = self._parse_jsonl(stream.read())
        elif suffix == ".json":
            sources = self._parse_json(stream.read())
        elif suffix == ".xml":
            sources = self._parse_xml(stream.read())
        elif suffix in {".html", ".htm"}:
            sources = self._parse_html(stream.read())
        elif suffix in {".ptpt", ".jcpt", ".stpt"}:
            sources = self._parse_zip_container_stream(stream)
        elif suffix == ".ppt":
            sources = self._parse_ppt(stream.read())
        elif suffix == ".lst":
            sources = self._parse_lst(file_name, stream.read())
        elif suffix == ".att":
            sources = self._parse_att(file_name, stream.read())
        elif suffix == ".gdb":
            sources = self._parse_gdb(file_name, stream.read())
        elif suffix == ".dll":
            sources = self._parse_binary(file_name, stream.read())
        elif not suffix:
            sources = self._parse_without_extension(file_name, stream.read())
        else:
            sources = [(None, None, _decode_text(stream.read()))]

        return self._chunk_sources(sources)

    def _parse_pdf(self, content: bytes) -> list[tuple[int | None, str | None, str]]:
        return self._parse_pdf_stream(BytesIO(content))

    def _parse_pdf_stream(self, stream: BinaryIO) -> list[tuple[int | None, str | None, str]]:
        from pypdf import PdfReader

        return [
            (index + 1, None, page.extract_text() or "")
            for index, page in enumerate(PdfReader(stream).pages)
        ]

    def _parse_docx(self, content: bytes) -> list[tuple[int | None, str | None, str]]:
        return self._parse_docx_stream(BytesIO(content))

    def _parse_docx_stream(self, stream: BinaryIO) -> list[tuple[int | None, str | None, str]]:
        from docx import Document

        try:
            document = Document(stream)
        except KeyError:
            stream.seek(0)
            content = stream.read()
            repaired_content = _repair_missing_optional_docx_parts(content)
            if repaired_content is content:
                raise
            document = Document(BytesIO(repaired_content))
        sources: list[tuple[int | None, str | None, str]] = []
        heading_stack: dict[int, str] = {}
        section_path: str | None = None
        parts: list[str] = []

        def flush() -> None:
            text = "\n".join(parts).strip()
            if text:
                sources.append((None, section_path, text))
            parts.clear()

        for block in document.iter_inner_content():
            if hasattr(block, "rows"):
                for row in block.rows:
                    values = [cell.text.strip() for cell in row.cells]
                    if any(values):
                        parts.append("\t".join(values))
                continue

            text = block.text.strip()
            if not text:
                continue
            heading_level = self._docx_heading_level(block)
            if heading_level:
                flush()
                heading_stack = {
                    level: heading
                    for level, heading in heading_stack.items()
                    if level < heading_level
                }
                heading_stack[heading_level] = text
                section_path = self._join_section_path(
                    *(heading_stack[level] for level in sorted(heading_stack))
                )
            parts.append(text)
        flush()
        return sources

    @staticmethod
    def _docx_heading_level(paragraph: Any) -> int | None:
        style = getattr(paragraph, "style", None)
        for value in (getattr(style, "style_id", ""), getattr(style, "name", "")):
            match = re.search(r"(?:heading|标题)\s*([1-6])$", str(value), re.IGNORECASE)
            if match:
                return int(match.group(1))
        return None

    def _parse_doc(self, content: bytes) -> list[tuple[int | None, str | None, str]]:
        if sys.platform != "win32":
            return self._parse_binary("legacy-word.doc", content)

        path = ""
        word = None
        document = None
        pythoncom = None
        import threading

        result: list[tuple[int | None, str | None, str]] = []
        error: Exception | None = None
        done = threading.Event()

        def _com_parse() -> None:
            nonlocal result, error, word, document, pythoncom
            try:
                import pythoncom
                from win32com.client import DispatchEx

                pythoncom.CoInitialize()
                with tempfile.NamedTemporaryFile(suffix=".doc", delete=False) as temporary:
                    temporary.write(content)
                    path = temporary.name

                word = DispatchEx("Word.Application")
                word.Visible = False
                word.DisplayAlerts = 0
                word.AutomationSecurity = 3
                document = word.Documents.Open(
                    path,
                    ConfirmConversions=False,
                    ReadOnly=True,
                    AddToRecentFiles=False,
                    OpenAndRepair=True,
                    NoEncodingDialog=True,
                )
                result = [(None, None, document.Content.Text)]
            except Exception as exc:
                error = exc
            finally:
                if document is not None:
                    try:
                        document.Close(False)
                    except Exception:
                        pass
                if word is not None:
                    try:
                        word.Quit()
                    except Exception:
                        pass
                if path:
                    try:
                        os.unlink(path)
                    except OSError:
                        pass
                if pythoncom is not None:
                    pythoncom.CoUninitialize()
                done.set()

        thread = threading.Thread(target=_com_parse, daemon=True)
        thread.start()
        if not done.wait(timeout=60):
            # COM 解析超时，回退到二进制提取
            return self._parse_binary("legacy-word.doc", content)
        if error or not result:
            return self._parse_binary("legacy-word.doc", content)
        return result

    def _parse_ppt(self, content: bytes) -> list[tuple[int | None, str | None, str]]:
        if sys.platform != "win32":
            return self._parse_binary("legacy-powerpoint.ppt", content)

        path = ""
        app = None
        presentation = None
        pythoncom = None
        import threading

        result: list[tuple[int | None, str | None, str]] = []
        error: Exception | None = None
        done = threading.Event()

        def _com_parse() -> None:
            nonlocal result, error, app, presentation, pythoncom
            try:
                import pythoncom
                from win32com.client import DispatchEx

                pythoncom.CoInitialize()
                with tempfile.NamedTemporaryFile(suffix=".ppt", delete=False) as temporary:
                    temporary.write(content)
                    path = temporary.name

                app = DispatchEx("PowerPoint.Application")
                app.Visible = False
                app.DisplayAlerts = 0
                presentation = app.Presentations.Open(path, WithWindow=False)
                parts: list[str] = []
                for slide in presentation.Slides:
                    for shape in slide.Shapes:
                        if shape.HasTextFrame and shape.TextFrame.HasText:
                            text = shape.TextFrame.TextRange.Text.strip()
                            if text:
                                parts.append(text)
                result = [(None, None, "\n".join(parts))]
            except Exception as exc:
                error = exc
            finally:
                if presentation is not None:
                    try:
                        presentation.Close()
                    except Exception:
                        pass
                if app is not None:
                    try:
                        app.Quit()
                    except Exception:
                        pass
                if path:
                    try:
                        os.unlink(path)
                    except OSError:
                        pass
                if pythoncom is not None:
                    pythoncom.CoUninitialize()
                done.set()

        thread = threading.Thread(target=_com_parse, daemon=True)
        thread.start()
        if not done.wait(timeout=60):
            return self._parse_binary("legacy-powerpoint.ppt", content)
        if error or not result:
            return self._parse_binary("legacy-powerpoint.ppt", content)
        return result

    def _parse_pptx(self, content: bytes) -> list[tuple[int | None, str | None, str]]:
        return self._parse_pptx_stream(BytesIO(content))

    def _parse_pptx_stream(self, stream: BinaryIO) -> list[tuple[int | None, str | None, str]]:
        try:
            with ZipFile(stream) as archive:
                sources = []
                for slide_number, slide_name in enumerate(self._pptx_slide_names(archive), start=1):
                    info = archive.getinfo(slide_name)
                    if info.file_size > PPTX_MAX_SLIDE_XML_BYTES:
                        raise ValueError(
                            f"PPTX 第 {slide_number} 页 XML 过大: {info.file_size // (1024 * 1024)} MB"
                        )
                    root = ElementTree.fromstring(archive.read(info))
                    parts = []
                    for paragraph in root.iter(f"{{{DRAWINGML_NAMESPACE}}}p"):
                        text_parts = []
                        for element in paragraph.iter():
                            if element.tag == f"{{{DRAWINGML_NAMESPACE}}}t" and element.text:
                                text_parts.append(element.text)
                            elif element.tag == f"{{{DRAWINGML_NAMESPACE}}}tab":
                                text_parts.append("\t")
                            elif element.tag == f"{{{DRAWINGML_NAMESPACE}}}br":
                                text_parts.append("\n")
                        text = "".join(text_parts).strip()
                        if text:
                            parts.append(text)
                    sources.append((slide_number, f"第 {slide_number} 页", "\n".join(parts)))
                return sources
        except BadZipFile as exc:
            raise ValueError("PPTX 文件损坏或格式不兼容") from exc

    @staticmethod
    def _pptx_slide_names(archive: ZipFile) -> list[str]:
        try:
            presentation = ElementTree.fromstring(archive.read("ppt/presentation.xml"))
            relationships = ElementTree.fromstring(archive.read("ppt/_rels/presentation.xml.rels"))
            targets = {
                relationship.get("Id"): posixpath.normpath(
                    posixpath.join("ppt", relationship.get("Target", "").replace("\\", "/"))
                )
                for relationship in relationships
                if relationship.tag == f"{{{PACKAGE_RELATIONSHIPS_NAMESPACE}}}Relationship"
                and relationship.get("TargetMode") != "External"
                and relationship.get("Type", "").endswith("/slide")
            }
            slide_names = []
            relationship_id_key = f"{{{OFFICE_RELATIONSHIPS_NAMESPACE}}}id"
            for slide_id in presentation.iter(f"{{{PRESENTATIONML_NAMESPACE}}}sldId"):
                target = targets.get(slide_id.get(relationship_id_key))
                if target and target in archive.NameToInfo:
                    slide_names.append(target)
            if slide_names:
                return slide_names
        except (KeyError, ElementTree.ParseError):
            pass

        def slide_number(name: str) -> int:
            match = re.search(r"slide(\d+)\.xml$", name)
            return int(match.group(1)) if match else sys.maxsize

        return sorted(
            (
                name
                for name in archive.namelist()
                if name.startswith("ppt/slides/slide") and name.endswith(".xml")
            ),
            key=slide_number,
        )

    def _parse_xlsx(self, content: bytes) -> list[tuple[int | None, str | None, str]]:
        return self._parse_xlsx_stream(BytesIO(content))

    def _parse_xlsx_stream(self, stream: BinaryIO) -> list[tuple[int | None, str | None, str]]:
        from openpyxl import load_workbook

        workbook = load_workbook(stream, read_only=True, data_only=True)
        try:
            sources = []
            for worksheet in workbook.worksheets:
                rows = self._tabular_rows(worksheet.iter_rows(values_only=True))
                sources.append((None, worksheet.title, "\n".join(rows)))
            return sources
        finally:
            workbook.close()

    def _parse_xls(self, content: bytes) -> list[tuple[int | None, str | None, str]]:
        import xlrd

        try:
            workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
        except Exception:
            sources = self._parse_xls_with_excel(content)
            return sources or self._parse_binary("legacy-excel.xls", content)
        try:
            sources = []
            for worksheet in workbook.sheets():
                rows = self._tabular_rows(
                    worksheet.row_values(row_index)
                    for row_index in range(worksheet.nrows)
                )
                sources.append((None, worksheet.name, "\n".join(rows)))
            return sources
        finally:
            workbook.release_resources()

    def _parse_xls_with_excel(self, content: bytes) -> list[tuple[int | None, str | None, str]]:
        if sys.platform != "win32":
            return []

        import threading

        result: list[tuple[int | None, str | None, str]] = []
        done = threading.Event()

        def _com_parse() -> None:
            path = ""
            excel = None
            workbook = None
            pythoncom = None
            try:
                import pythoncom
                from win32com.client import DispatchEx

                pythoncom.CoInitialize()
                with tempfile.NamedTemporaryFile(suffix=".xls", delete=False) as temporary:
                    temporary.write(content)
                    path = temporary.name

                excel = DispatchEx("Excel.Application")
                excel.Visible = False
                excel.DisplayAlerts = False
                excel.AutomationSecurity = 3
                workbook = excel.Workbooks.Open(
                    path,
                    UpdateLinks=0,
                    ReadOnly=True,
                    IgnoreReadOnlyRecommended=True,
                    AddToMru=False,
                    Notify=False,
                )
                for worksheet in workbook.Worksheets:
                    values = worksheet.UsedRange.Value2
                    if values is None:
                        rows = []
                    elif isinstance(values, tuple) and values and isinstance(values[0], tuple):
                        rows = values
                    elif isinstance(values, tuple):
                        rows = [values]
                    else:
                        rows = [(values,)]
                    result.append((None, str(worksheet.Name), "\n".join(self._tabular_rows(rows))))
            except Exception:
                result.clear()
            finally:
                if workbook is not None:
                    try:
                        workbook.Close(False)
                    except Exception:
                        pass
                if excel is not None:
                    try:
                        excel.Quit()
                    except Exception:
                        pass
                if path:
                    try:
                        os.unlink(path)
                    except OSError:
                        pass
                if pythoncom is not None:
                    pythoncom.CoUninitialize()
                done.set()

        thread = threading.Thread(target=_com_parse, daemon=True)
        thread.start()
        if not done.wait(timeout=60):
            return []
        return result

    def _parse_csv(self, content: bytes) -> list[tuple[int | None, str | None, str]]:
        text = _decode_text(content)
        sample = text[:8192]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        except csv.Error:
            dialect = csv.excel
        rows = self._tabular_rows(csv.reader(StringIO(text), dialect))
        return [(None, None, "\n".join(rows))]

    def _parse_json(self, content: bytes) -> list[tuple[int | None, str | None, str]]:
        value = json.loads(_decode_text(content))
        return [(None, None, json.dumps(value, ensure_ascii=False, indent=2, default=str))]

    def _parse_jsonl(self, content: bytes) -> list[tuple[int | None, str | None, str]]:
        records = []
        for line_number, line in enumerate(_decode_text(content).splitlines(), start=1):
            if not line.strip():
                continue
            try:
                value = json.loads(line)
            except json.JSONDecodeError as exc:
                raise ValueError(f"JSONL 第 {line_number} 行格式错误") from exc
            records.append(json.dumps(value, ensure_ascii=False, default=str))
        return [(None, None, "\n".join(records))]

    def _parse_xml(self, content: bytes) -> list[tuple[int | None, str | None, str]]:
        root = ElementTree.fromstring(content)
        lines = []
        seen = set()
        for element in root.iter():
            values = [element.text, *element.attrib.values()]
            for value in values:
                if not value:
                    continue
                text = str(value).replace("\\r\\n", "\n").strip()
                if text and text not in seen:
                    seen.add(text)
                    lines.append(text)
        return [(None, root.tag, "\n".join(lines))]

    def _parse_html(self, content: bytes) -> list[tuple[int | None, str | None, str]]:
        extractor = _TextExtractor()
        extractor.feed(_decode_text(content))
        return [(None, None, "\n".join(extractor.parts))]

    def _parse_zip_container(self, content: bytes) -> list[tuple[int | None, str | None, str]]:
        return self._parse_zip_container_stream(BytesIO(content))

    def _parse_zip_container_stream(self, stream: BinaryIO) -> list[tuple[int | None, str | None, str]]:
        sources = []
        total_size = 0
        try:
            with ZipFile(stream) as archive:
                for entry in archive.infolist()[:200]:
                    if entry.is_dir() or entry.file_size > 20 * 1024 * 1024:
                        continue
                    total_size += entry.file_size
                    if total_size > 50 * 1024 * 1024:
                        break
                    suffix = Path(entry.filename).suffix.lower()
                    if suffix not in {".xml", ".json", ".csv", ".txt", ".md", ".html", ".htm"}:
                        continue
                    data = archive.read(entry)
                    if suffix == ".xml":
                        nested = self._parse_xml(data)
                    elif suffix == ".json":
                        nested = self._parse_json(data)
                    elif suffix == ".csv":
                        nested = self._parse_csv(data)
                    elif suffix in {".html", ".htm"}:
                        nested = self._parse_html(data)
                    else:
                        nested = [(None, None, _decode_text(data))]
                    sources.extend(
                        (page_number, entry.filename, text)
                        for page_number, _section, text in nested
                    )
        except BadZipFile as exc:
            raise ValueError("压缩容器损坏或格式不兼容") from exc
        return sources

    def _parse_without_extension(self, file_name: str, content: bytes) -> list[tuple[int | None, str | None, str]]:
        if self._looks_like_text(content):
            return [(None, None, _decode_text(content))]
        if content.startswith(b"PK\x03\x04"):
            return self._parse_zip_container(content)
        return self._parse_binary(file_name, content)

    def _parse_att(self, file_name: str, content: bytes) -> list[tuple[int | None, str | None, str]]:
        if content.startswith(GEOMAP_LAYER_STYLE_MAGIC):
            return self._parse_geomap_layer_style(file_name, content)
        return self._parse_without_extension(file_name, content)

    def _parse_gdb(self, file_name: str, content: bytes) -> list[tuple[int | None, str | None, str]]:
        if content.startswith(GEOMAP_MAP_MAGIC):
            return self._parse_geomap_map(file_name, content)
        return self._parse_binary(file_name, content)

    def _parse_lst(self, file_name: str, content: bytes) -> list[tuple[int | None, str | None, str]]:
        if content.startswith(GEOMAP_ALBUM_MAGIC):
            return self._parse_geomap_album(file_name, content)
        return [(None, None, _decode_text(content))]

    def _parse_geomap_album(
        self,
        file_name: str,
        content: bytes,
    ) -> list[tuple[int | None, str | None, str]]:
        if len(content) <= GEOMAP_ALBUM_TREE_OFFSET:
            raise ValueError("GeoMap Album Information 文件损坏：缺少图册树数据")

        cursor = GEOMAP_ALBUM_TREE_OFFSET

        def read_u32() -> int:
            nonlocal cursor
            if cursor + 4 > len(content):
                raise ValueError("读取 uint32 时超出文件边界")
            value = struct.unpack_from("<I", content, cursor)[0]
            cursor += 4
            return value

        def read_string() -> str:
            nonlocal cursor
            size = read_u32()
            if size > 1024 * 1024 or cursor + size > len(content):
                raise ValueError(f"非法的 GBK 字符串长度: {size}")
            raw = content[cursor:cursor + size]
            cursor += size
            try:
                return raw.decode("gb18030")
            except UnicodeDecodeError as exc:
                raise ValueError("GeoMap 图册字符串不是有效的 GBK/GB18030") from exc

        try:
            stored_file_name = read_string()
            flags = read_u32()
            album_name = read_string()
            category_count = read_u32()
            if category_count > 10_000:
                raise ValueError(f"非法的图册分类数: {category_count}")

            categories: list[tuple[str, list[tuple[str, str, str]]]] = []
            total_items = 0
            for _ in range(category_count):
                node_type = read_u32()
                if node_type != 3:
                    raise ValueError(f"未知的图册分类节点类型: {node_type}")
                category_name = read_string()
                item_count = read_u32()
                if item_count > 100_000:
                    raise ValueError(f"非法的图件数: {item_count}")
                items = []
                for _ in range(item_count):
                    item_type = read_u32()
                    if item_type != 1:
                        raise ValueError(f"未知的图册图件节点类型: {item_type}")
                    display_name = read_string()
                    folder_path = read_string().replace("\\", "/").strip("/")
                    database_name = read_string()
                    items.append((display_name, folder_path, database_name))
                total_items += len(items)
                categories.append((category_name, items))

            if cursor != len(content):
                raise ValueError(f"图册树结束后仍有 {len(content) - cursor} 字节未解析数据")
        except ValueError as exc:
            raise ValueError(f"GeoMap Album Information 文件损坏: {exc}") from exc

        sources: list[tuple[int | None, str | None, str]] = [
            (
                None,
                "GeoMap Album Information",
                "\n".join(
                    (
                        f"文件名: {file_name}",
                        "格式: GeoMap Album Information 图册信息文件",
                        "字符编码: GBK/GB18030",
                        f"图册名: {album_name or stored_file_name}",
                        f"分类数: {len(categories)}",
                        f"图件数: {total_items}",
                        f"图册标志: {flags}",
                    )
                ),
            )
        ]
        for category_name, items in categories:
            for display_name, folder_path, database_name in items:
                relative_path = f"{folder_path}/{database_name}" if folder_path else database_name
                sources.append(
                    (
                        None,
                        f"GeoMap图册/{category_name}/{display_name}",
                        "\n".join(
                            (
                                f"图册: {album_name or stored_file_name}",
                                f"分类: {category_name}",
                                f"图件显示名: {display_name}",
                                f"相对目录: {folder_path or '.'}",
                                f"GeoMap 数据库文件: {database_name}",
                                f"完整相对路径: {relative_path}",
                            )
                        ),
                    )
                )
        return sources

    def _parse_geomap_layer_style(
        self,
        file_name: str,
        content: bytes,
    ) -> list[tuple[int | None, str | None, str]]:
        body_size = len(content) - GEOMAP_LAYER_STYLE_HEADER_SIZE
        if body_size <= 0 or body_size % GEOMAP_LAYER_STYLE_RECORD_SIZE:
            raise ValueError(
                "Geomap v3.60 LayerStyle 文件损坏："
                f"文件头后的数据长度不是 {GEOMAP_LAYER_STYLE_RECORD_SIZE} 字节记录的整数倍"
            )

        record_count = body_size // GEOMAP_LAYER_STYLE_RECORD_SIZE
        declared_record_count = struct.unpack_from("<I", content, 512)[0]
        if declared_record_count not in {0, record_count}:
            raise ValueError(
                "Geomap v3.60 LayerStyle 文件损坏："
                f"文件头记录数为 {declared_record_count}，实际记录数为 {record_count}"
            )

        table_rows = ["记录号\t图层或属性组\t属性字段1\t属性字段2\t属性字段3\t属性字段4\t属性字段5\t属性字段6"]

        for index in range(record_count):
            start = GEOMAP_LAYER_STYLE_HEADER_SIZE + index * GEOMAP_LAYER_STYLE_RECORD_SIZE
            record = content[start:start + GEOMAP_LAYER_STYLE_RECORD_SIZE]
            fields = [
                self._decode_geomap_text_slot(record[offset:offset + width])
                for offset, width in GEOMAP_LAYER_STYLE_TEXT_SLOTS
            ]
            while fields and not fields[-1]:
                fields.pop()
            if not fields:
                fields = [f"图层_{index + 1}"]
            table_rows.append("\t".join((str(index + 1), *fields)))

        return [
            (
                None,
                "GeoMap属性表/LayerStyle",
                "\n".join(
                    (
                        f"文件名: {file_name}",
                        "格式: Geomap v3.60 LayerStyle 属性数据表",
                        "字符编码: GBK/GB18030",
                        f"LayerStyle 记录数: {record_count}",
                        f"固定记录长度: {GEOMAP_LAYER_STYLE_RECORD_SIZE} 字节",
                        "说明: 每行对应同序号 GDB 图层的样式或属性字段定义。",
                        "属性数据表:",
                        *table_rows,
                    )
                ),
            )
        ]

    def _parse_geomap_map(
        self,
        file_name: str,
        content: bytes,
    ) -> list[tuple[int | None, str | None, str]]:
        layer_offsets = [match.start() for match in re.finditer(re.escape(GEOMAP_LAYER_MAGIC), content)]
        if not layer_offsets:
            raise ValueError("Geomap v3.60 Map 文件损坏：没有找到图层数据")

        map_name = self._decode_geomap_text_slot(content[0x204:0x304]) or Path(file_name).stem
        layers: list[tuple[str, list[str]]] = []
        total_attributes = 0

        for index, layer_offset in enumerate(layer_offsets):
            layer_end = layer_offsets[index + 1] if index + 1 < len(layer_offsets) else len(content)
            layer_size = layer_end - layer_offset
            minimum_named_header_size = GEOMAP_LAYER_NAME_OFFSET
            if layer_size < minimum_named_header_size:
                raise ValueError(
                    f"Geomap v3.60 Map 文件损坏：第 {index + 1} 个图层缺少名称字段"
                )

            name_length = struct.unpack_from(
                "<I",
                content,
                layer_offset + GEOMAP_LAYER_NAME_LENGTH_OFFSET,
            )[0]
            if not 0 < name_length <= 256:
                raise ValueError(
                    f"Geomap v3.60 Map 文件损坏：第 {index + 1} 个图层名称长度非法: {name_length}"
                )
            name_start = layer_offset + GEOMAP_LAYER_NAME_OFFSET
            name_end = name_start + name_length
            if name_end > min(layer_offset + GEOMAP_LAYER_HEADER_SIZE, layer_end):
                raise ValueError(f"Geomap v3.60 Map 文件损坏：第 {index + 1} 个图层名称越界")
            try:
                layer_name = content[name_start:name_end].decode("gb18030").strip()
            except UnicodeDecodeError as exc:
                raise ValueError(
                    f"Geomap v3.60 Map 文件损坏：第 {index + 1} 个图层名称不是有效的 GBK/GB18030"
                ) from exc
            if not layer_name:
                layer_name = f"图层_{index + 1}"

            is_final_layer = index + 1 == len(layer_offsets)
            if layer_size < GEOMAP_LAYER_HEADER_SIZE and not is_final_layer:
                raise ValueError(
                    f"Geomap v3.60 Map 文件损坏：第 {index + 1} 个中间图层短于 "
                    f"{GEOMAP_LAYER_HEADER_SIZE} 字节"
                )

            attributes = self._extract_geomap_layer_attributes(
                content,
                min(layer_offset + GEOMAP_LAYER_HEADER_SIZE, layer_end),
                layer_end,
            )
            total_attributes += len(attributes)
            if total_attributes > GEOMAP_MAX_ATTRIBUTE_ROWS:
                raise ValueError(
                    f"Geomap v3.60 Map 可提取属性超过 {GEOMAP_MAX_ATTRIBUTE_ROWS} 行，已停止解析"
                )
            layers.append((layer_name, attributes))

        sources: list[tuple[int | None, str | None, str]] = [
            (
                None,
                "GeoMap属性表/概览",
                "\n".join(
                    (
                        f"文件名: {file_name}",
                        "格式: Geomap v3.60 Map 属性数据表",
                        "字符编码: GBK/GB18030",
                        f"图名: {map_name}",
                        f"图层数: {len(layers)}",
                        f"提取的文本或数值属性记录数: {total_attributes}",
                        "说明: 仅提取图层名称及对象中的文本、数值属性；图形坐标和二进制样式不作为文本入库。",
                    )
                ),
            )
        ]
        for index, (layer_name, attributes) in enumerate(layers, start=1):
            rows = ["记录号\t属性值"]
            rows.extend(f"{row_index}\t{value}" for row_index, value in enumerate(attributes, start=1))
            if not attributes:
                rows.append("-\t本图层没有可直接提取的文本或数值属性")
            sources.append(
                (
                    None,
                    f"GeoMap属性表/{index:03d}/{layer_name}",
                    "\n".join(
                        (
                            "GeoMap 图层属性数据表",
                            f"图层序号: {index}/{len(layers)}",
                            f"图层名称: {layer_name}",
                            f"属性记录数: {len(attributes)}",
                            *rows,
                        )
                    ),
                )
            )
        return sources

    def _extract_geomap_layer_attributes(
        self,
        content: bytes,
        body_start: int,
        body_end: int,
    ) -> list[str]:
        high_confidence_offsets: set[int] = set()
        marker_offset = content.find(GEOMAP_OBJECT_TEXT_MARKER, body_start, body_end)
        while marker_offset >= 0:
            high_confidence_offsets.update((marker_offset + 4, marker_offset + 12))
            marker_offset = content.find(
                GEOMAP_OBJECT_TEXT_MARKER,
                marker_offset + len(GEOMAP_OBJECT_TEXT_MARKER),
                body_end,
            )

        candidates: list[tuple[int, int, str, bool]] = []
        offset = body_start
        while offset + 4 <= body_end:
            value = self._decode_geomap_length_prefixed_text(content, offset, body_end)
            if value is None:
                offset += 1
                continue
            byte_length, text = value
            high_confidence = offset in high_confidence_offsets
            if high_confidence or self._looks_like_geomap_attribute(text):
                candidates.append((offset + 4, offset + 4 + byte_length, text, high_confidence))
                offset += 4 + byte_length
            else:
                offset += 1

        # Marker-backed fields win if a looser scan found an overlapping binary lookalike.
        candidates.sort(key=lambda item: (item[0], item[1]))
        selected: list[tuple[int, int, str, bool]] = []
        for candidate in candidates:
            start, _end, _text, high_confidence = candidate
            if selected and start < selected[-1][1]:
                if high_confidence and not selected[-1][3]:
                    selected[-1] = candidate
                continue
            selected.append(candidate)
        return [text for _start, _end, text, _high_confidence in selected]

    @staticmethod
    def _decode_geomap_length_prefixed_text(
        content: bytes,
        length_offset: int,
        end: int,
    ) -> tuple[int, str] | None:
        byte_length = struct.unpack_from("<I", content, length_offset)[0]
        if not 0 < byte_length <= GEOMAP_MAX_TEXT_FIELD_BYTES:
            return None
        value_start = length_offset + 4
        value_end = value_start + byte_length
        if value_end > end:
            return None
        raw = content[value_start:value_end]
        if any(byte < 32 for byte in raw):
            return None
        try:
            value = raw.decode("gb18030")
        except UnicodeDecodeError:
            return None
        value = re.sub(r"\s+", " ", value).strip()
        if not value or any(not character.isprintable() for character in value):
            return None
        meaningful = sum(
            character.isalnum()
            or "\u4e00" <= character <= "\u9fff"
            or character in " ._-/|#()（）+%"
            for character in value
        )
        if meaningful / len(value) < 0.85:
            return None
        return byte_length, value

    @staticmethod
    def _looks_like_geomap_attribute(value: str) -> bool:
        chinese_count = sum("\u4e00" <= character <= "\u9fff" for character in value)
        if chinese_count >= 2:
            return True
        if re.fullmatch(r"[-+]?\d+(?:\.\d+)?(?:[|/]\d+)*", value):
            return True
        return bool(re.fullmatch(r"[A-Za-z][A-Za-z0-9_.-]{2,}", value))

    @staticmethod
    def _decode_geomap_text_slot(slot: bytes) -> str:
        raw = slot.split(b"\x00", 1)[0]
        if not raw:
            return ""
        try:
            value = raw.decode("gb18030").strip()
        except UnicodeDecodeError:
            return ""
        if not value or any(ord(character) < 32 for character in value):
            return ""
        meaningful = sum(
            character.isalnum()
            or "\u4e00" <= character <= "\u9fff"
            or character in "._- /"
            for character in value
        )
        return value if meaningful / len(value) >= 0.7 else ""

    def _looks_like_text(self, content: bytes) -> bool:
        sample = content[:8192]
        if not sample:
            return False
        if sample.startswith((b"\xff\xfe", b"\xfe\xff")):
            return True
        zero_ratio = sample.count(0) / len(sample)
        control_count = sum(byte < 9 or 13 < byte < 32 for byte in sample)
        return zero_ratio < 0.02 and control_count / len(sample) < 0.05

    def _parse_binary(self, file_name: str, content: bytes) -> list[tuple[int | None, str | None, str]]:
        strings: list[str] = []
        seen = set()
        total_characters = 0

        def append(value: str) -> None:
            nonlocal total_characters
            value = re.sub(r"\s+", " ", value).strip()
            if len(value) < 4 or value in seen:
                return
            meaningful = sum(character.isalnum() or "\u4e00" <= character <= "\u9fff" for character in value)
            if meaningful / len(value) < 0.35:
                return
            seen.add(value)
            strings.append(value)
            total_characters += len(value)

        for match in re.finditer(rb"[\x20-\x7e]{4,}", content):
            append(match.group().decode("ascii"))
            if total_characters >= 100_000:
                break

        if total_characters < 100_000:
            for match in re.finditer(rb"(?:[\x20-\x7e]\x00){4,}", content):
                append(match.group().decode("utf-16-le"))
                if total_characters >= 100_000:
                    break

        if total_characters < 100_000:
            for match in re.finditer(rb"(?:[\x20-\x7e]|[\x81-\xfe][\x40-\xfe]){4,}", content):
                try:
                    append(match.group().decode("gb18030"))
                except UnicodeDecodeError:
                    continue
                if total_characters >= 100_000:
                    break

        header = f"文件名: {file_name}\n文件大小: {len(content)} 字节\n可读字符串:"
        return [(None, "二进制内容", f"{header}\n" + "\n".join(strings))]

    def _tabular_rows(self, rows: Iterable[Iterable[Any]]) -> list[str]:
        result = []
        for row in rows:
            values = [_cell_text(value) for value in row]
            while values and not values[-1]:
                values.pop()
            if any(values):
                result.append("\t".join(values))
        return result

    def _chunk_sources(
        self,
        sources: Iterable[tuple[int | None, str | None, str]],
    ) -> list[ParsedChunk]:
        chunks: list[ParsedChunk] = []
        for page_number, source_section_path, raw_text in sources:
            text = self._normalize_chunk_text(raw_text)
            if not text:
                continue

            for section_path, section_text in self._split_source_sections(source_section_path, text):
                for chunk_text in self._split_section_text(section_text):
                    chunks.append(
                        ParsedChunk(
                            index=len(chunks),
                            text=chunk_text,
                            page_number=page_number,
                            section_path=section_path,
                        )
                    )
        return chunks

    @staticmethod
    def _normalize_chunk_text(raw_text: str) -> str:
        text = raw_text.replace("\x00", "").replace("\r\n", "\n").replace("\r", "\n")
        text = re.sub(r"[ \t]+", " ", text)
        text = re.sub(r" *\n *", "\n", text)
        return re.sub(r"\n{3,}", "\n\n", text).strip()

    def _split_source_sections(
        self,
        source_section_path: str | None,
        text: str,
    ) -> list[tuple[str | None, str]]:
        sections: list[tuple[str | None, str]] = []
        heading_stack: dict[int, str] = {}
        current_path = source_section_path
        current_lines: list[str] = []
        fence_marker: str | None = None

        def flush() -> None:
            section_text = "\n".join(current_lines).strip()
            if section_text:
                sections.append((current_path, section_text))
            current_lines.clear()

        for line in text.splitlines():
            stripped = line.strip()
            marker_match = re.match(r"^(?:\x60{3,}|~{3,})", stripped)
            if marker_match:
                marker = marker_match.group(0)[0]
                if fence_marker == marker:
                    fence_marker = None
                elif fence_marker is None:
                    fence_marker = marker

            heading = None if fence_marker else self._heading_details(stripped)
            if heading:
                flush()
                level, title = heading
                heading_stack = {
                    existing_level: existing_title
                    for existing_level, existing_title in heading_stack.items()
                    if existing_level < level
                }
                heading_stack[level] = title
                current_path = self._join_section_path(
                    source_section_path,
                    *(heading_stack[key] for key in sorted(heading_stack)),
                )
            current_lines.append(line)

        flush()
        return sections

    @staticmethod
    def _heading_details(line: str) -> tuple[int, str] | None:
        if not line or len(line) > 120 or line.endswith(("。", "！", "？", "!", "?", "；", ";")):
            return None

        markdown_match = MARKDOWN_HEADING_PATTERN.match(line)
        if markdown_match:
            return len(markdown_match.group(1)), markdown_match.group(2).strip()

        for pattern, fixed_level in NUMBERED_HEADING_PATTERNS:
            match = pattern.match(line)
            if not match:
                continue
            if fixed_level is None:
                number = match.group(1)
                title = match.group(2).strip()
                return min(number.count(".") + 1, 6), f"{number} {title}"
            title = match.group(1).strip()
            return fixed_level, line if line.startswith("第") else title
        return None

    @staticmethod
    def _join_section_path(*parts: str | None) -> str | None:
        cleaned: list[str] = []
        for part in parts:
            for segment in part.split("/") if part else ():
                value = segment.strip(" \t")
                if value and (not cleaned or cleaned[-1] != value):
                    cleaned.append(value)
        return "/".join(cleaned) or None

    def _split_section_text(self, text: str) -> list[str]:
        chunks: list[str] = []
        start = 0
        while start < len(text):
            hard_end = min(start + CHUNK_SIZE, len(text))
            end = self._natural_chunk_end(text, start, hard_end)
            chunk_text = text[start:end].strip()
            if chunk_text:
                chunks.append(chunk_text)
            if end >= len(text):
                break
            next_start = self._natural_overlap_start(text, start, end)
            start = next_start if next_start > start else end
        return chunks

    @staticmethod
    def _natural_chunk_end(text: str, start: int, hard_end: int) -> int:
        if hard_end >= len(text):
            return len(text)

        search_start = min(start + MIN_NATURAL_CHUNK_SIZE, hard_end)
        window = text[search_start:hard_end]
        boundary_patterns = (
            re.compile(r"\n\n"),
            re.compile(r"\n"),
            re.compile(r"[。！？!?；;](?:[\"'”’）)])?"),
            re.compile(r"\s+"),
        )
        for pattern in boundary_patterns:
            matches = list(pattern.finditer(window))
            if matches:
                return search_start + matches[-1].end()
        return hard_end

    @staticmethod
    def _natural_overlap_start(text: str, chunk_start: int, chunk_end: int) -> int:
        target = max(chunk_start + 1, chunk_end - CHUNK_OVERLAP)
        window = text[target:chunk_end]
        boundary = re.search(r"(?:\n+|[。！？!?；;](?:[\"'”’）)])?\s*|\s+)", window)
        if boundary:
            return target + boundary.end()
        return target
